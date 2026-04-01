import json
import os
import re
import shutil
import subprocess
import tempfile
from pathlib import Path

from openpyxl import load_workbook

XLSX_PATH = Path(os.environ.get("ETQ_XLSX_PATH", "specs/ETQ.xlsx"))
FIELD_REGISTRY_PATH = Path("src/model/fieldRegistry.ts")
OUT_PATH = Path("/tmp/etq_live_extract.json")


def read_ages_and_series(ws, row: int) -> list[float]:
    values: list[float] = []
    col = 3  # C
    while True:
        v = ws.cell(row=row, column=col).value
        if v is None or v == "":
            break
        values.append(float(v))
        col += 1
    return values


def load_input_cells() -> list[str]:
    field_registry = FIELD_REGISTRY_PATH.read_text(encoding="utf-8")
    return re.findall(r'^\s*(B\d+):\s+"[^"]+",?$', field_registry, flags=re.MULTILINE)


def load_raw_inputs(inputs_ws) -> dict:
    raw: dict = {}
    for cell in load_input_cells():
        raw[cell] = inputs_ws[cell].value
    return raw


def recalc_workbook(source_path: Path) -> Path:
    tmpdir = Path(tempfile.mkdtemp(prefix="etq-live-"))
    in_dir = tmpdir / "in"
    out_dir = tmpdir / "out"
    in_dir.mkdir(parents=True, exist_ok=True)
    out_dir.mkdir(parents=True, exist_ok=True)
    source_copy = in_dir / source_path.name
    shutil.copy2(source_path, source_copy)
    subprocess.run(
        [
            "soffice",
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(out_dir),
            str(source_copy),
        ],
        check=True,
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
    )
    return out_dir / source_path.name


def main() -> None:
    wb = load_workbook(recalc_workbook(XLSX_PATH), data_only=True)
    ws_inputs = wb["Inputs"]
    ws_norm = wb["RetNorm_Engine"]
    ws_early = wb["RetEarly_Engine"]
    month_remaining = ws_inputs["B302"].value
    projection_month_override = None
    if isinstance(month_remaining, (int, float)):
        projection_month_override = int(round(13 - float(month_remaining)))

    payload = {
        "early": int(float(ws_early["B3"].value)),
        "projection_month_override": projection_month_override,
        "raw": load_raw_inputs(ws_inputs),
        "exp": {
            "years": read_ages_and_series(ws_norm, 2),
            "ages": read_ages_and_series(ws_norm, 3),
            "cashE": read_ages_and_series(ws_early, 381),
            "cashN": read_ages_and_series(ws_norm, 381),
            "nwE": read_ages_and_series(ws_early, 382),
            "nwN": read_ages_and_series(ws_norm, 382),
        },
    }

    OUT_PATH.write_text(json.dumps(payload), encoding="utf-8")
    print(f"wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
